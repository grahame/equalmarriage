#!/usr/bin/env python3

import openpyxl
from collections import defaultdict
import json


def dump(fname, obj):
    with open(fname, 'w') as fd:
        json.dump(obj, fd, indent=4, separators=(',', ': '), sort_keys=True)


def fix_val(v):
    return v


def brute_read(sheet, offset):
    return [t for t in [[fix_val(s.value) for s in t if s.value] for t in sheet.iter_rows(row_offset=offset)] if len(t) > 0]


def fix_electorate(s):
    return s.rsplit('(', 1)[0].strip()


def scrape_participation_aspect(electorate_obj, state_obj, sheet, aspect):
    rows = brute_read(sheet, 5)
    header = rows[0]
    rows = rows[1:]

    current_state = None
    target = None
    for row in rows:
        if len(row) == 1:
            if row[0].endswith(' Divisions'):
                current_state = row[0].rsplit(' ', 1)[0]
                continue
            else:
                break
        if row[1] == 'Total participants':
            current_electorate = row[0]
            if current_electorate.endswith('(Total)'):
                target = "state"
            else:
                target = "electorate"
            current_electorate = fix_electorate(current_electorate)
            row = row[1:]

        attr = row[0]
        if target == "electorate":
            target_dict = electorate_obj[current_state][current_electorate][aspect]
        else:
            target_dict = state_obj[current_state][aspect]
        for k, v in zip(header, row[1:]):
            target_dict[k][attr] = v


def scrape_participation(fname):
    wb = openpyxl.load_workbook(fname)
    electorate_obj = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(dict))))
    state_obj = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(dict))))
    scrape_participation_aspect(electorate_obj, state_obj, wb.get_sheet_by_name("Table 4"), "All")
    scrape_participation_aspect(electorate_obj, state_obj, wb.get_sheet_by_name("Table 5"), "Male")
    scrape_participation_aspect(electorate_obj, state_obj, wb.get_sheet_by_name("Table 6"), "Female")
    dump('output/participation/By State.json', state_obj)
    dump('output/participation/By Electorate.json', electorate_obj)


def scrape_response(fname):
    wb = openpyxl.load_workbook(fname)
    sheet = wb.get_sheet_by_name('Table 2')
    rows = brute_read(sheet, 6)
    header = rows[0]
    rows = rows[1:]

    obj = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(dict))))

    current_state = None
    header = []
    topics = ['Yes', 'No', 'Total', 'Response clear', 'Response not clear', 'Non-responding', 'Total']
    for topic in topics:
        for aspect in ['count', '%']:
            header.append((topic, aspect))
    for row in rows:
        if len(row) == 1:
            if row[0].endswith(' Divisions'):
                current_state = row[0].rsplit(' ', 1)[0]
                continue
            else:
                break
        row_type = 'Electorate'
        if row[0].endswith('(Total)'):
            row_type = 'State'
        target = fix_electorate(row[0])

        for (topic, aspect), value in zip(header, row[1:]):
            if row_type == 'Electorate':
                obj[row_type][current_state][target][topic][aspect] = value
            else:
                obj[row_type][target][topic][aspect] = value

    def dump_elem(elem):
        dump('output/response/By %s.json' % (elem), obj[elem])

    dump_elem('State')
    dump_elem('Electorate')


def main():
    scrape_participation('data/xlsx/participation.xlsx')
    scrape_response('data/xlsx/response.xlsx')
    pass


if __name__ == '__main__':
    main()
